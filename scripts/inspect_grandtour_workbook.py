from openpyxl import load_workbook
import sys
path = r"C:\Users\Tony\OneDrive\Desktop\grandtour_dummy_users_and_tips.xlsx"
print('Workbook path:', path)
try:
    wb = load_workbook(path, read_only=True)
except Exception as e:
    print('ERROR opening workbook:', e)
    sys.exit(2)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n=== Sheet:', name, '===')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(i, row)
        if i >= 10:
            break
